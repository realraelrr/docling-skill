"""Spreadsheet ingestion helpers."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from docling.datamodel.base_models import InputFormat
from docling.document_converter import CsvFormatOption, ExcelFormatOption


def _extract_spreadsheet_metadata(
    structured_document: dict[str, Any],
    *,
    source_format: str | None = None,
    normalized_from: str | None = None,
) -> dict[str, Any]:
    tables = structured_document.get("tables", [])
    groups = structured_document.get("groups", [])
    pages = structured_document.get("pages", {})
    sheet_names = [
        group.get("name")
        for group in groups
        if isinstance(group, dict)
        and isinstance(group.get("name"), str)
        and group["name"].startswith("sheet:")
    ]
    sheet_count = len(pages) if isinstance(pages, dict) else len(pages or [])
    if sheet_count == 0:
        sheet_count = len(sheet_names)
    if sheet_count == 0 and tables:
        sheet_count = 1

    merged_cell_count = 0
    for table in tables:
        if not isinstance(table, dict):
            continue
        data = table.get("data", {})
        if not isinstance(data, dict):
            continue
        for cell in data.get("table_cells", []):
            if not isinstance(cell, dict):
                continue
            if cell.get("row_span", 1) > 1 or cell.get("col_span", 1) > 1:
                merged_cell_count += 1

    metadata = {
        "sheet_count": sheet_count,
        "table_count": len(tables),
        "merged_cell_count": merged_cell_count,
        "has_merged_cells": merged_cell_count > 0,
        "has_multi_sheet": sheet_count > 1,
    }
    if source_format is not None:
        metadata["source_format"] = source_format
    if normalized_from is not None:
        metadata["normalized_from"] = normalized_from
    return metadata


def _spreadsheet_format_option(input_format: InputFormat):
    if input_format == InputFormat.CSV:
        return CsvFormatOption()
    return ExcelFormatOption()


def _safe_excel_sheet_title(title: str, fallback: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", title).strip() or fallback
    return cleaned[:31]


def _xls_cell_value(book: Any, cell: Any) -> Any:
    import xlrd

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, book.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.biffh.error_text_from_code.get(cell.value, f"#ERR{cell.value}")
    return cell.value


def _normalize_xls_to_xlsx(input_path: Path, output_path: Path) -> Path:
    try:
        import openpyxl
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "XLS support requires xlrd and openpyxl. Save as .xlsx or .csv before ingestion."
        ) from exc

    try:
        workbook = xlrd.open_workbook(str(input_path), formatting_info=True)
    except Exception as exc:
        raise RuntimeError(
            "Unable to read XLS file. Save as .xlsx or .csv before ingestion."
        ) from exc

    normalized_workbook = openpyxl.Workbook()
    normalized_workbook.remove(normalized_workbook.active)

    for sheet_index, sheet in enumerate(workbook.sheets(), start=1):
        worksheet = normalized_workbook.create_sheet(
            title=_safe_excel_sheet_title(sheet.name, f"Sheet{sheet_index}")
        )
        for row_index in range(sheet.nrows):
            for column_index in range(sheet.ncols):
                value = _xls_cell_value(workbook, sheet.cell(row_index, column_index))
                if value in {None, ""}:
                    continue
                worksheet.cell(
                    row=row_index + 1,
                    column=column_index + 1,
                    value=value,
                )
        for row_start, row_end, column_start, column_end in sheet.merged_cells:
            worksheet.merge_cells(
                start_row=row_start + 1,
                end_row=row_end,
                start_column=column_start + 1,
                end_column=column_end,
            )

    if not normalized_workbook.worksheets:
        normalized_workbook.create_sheet(title="Sheet1")

    normalized_workbook.save(output_path)
    return output_path
